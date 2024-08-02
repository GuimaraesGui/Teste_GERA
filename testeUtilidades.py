import pandas as pd
from openpyxl import load_workbook

def preencher_celulas_vazias():
    caminho_arquivo_base_dados = "C:\\Users\\00805129\\OneDrive - NATURGY INFORMATICA S.A\\Escritorio\\testeUtlidades\\Teste.xlsx"
    caminho_arquivo_atual = "C:\\Users\\00805129\\OneDrive - NATURGY INFORMATICA S.A\\Escritorio\\testeUtlidades\\Teste Utilidades.xlsm"
   
    df_base_dados = pd.read_excel(caminho_arquivo_base_dados, sheet_name="Worksheet", usecols="A:K")
    
    df_atual = pd.read_excel(caminho_arquivo_atual, sheet_name="Geral", usecols="A:BB")
    
    wb_atual = load_workbook(caminho_arquivo_atual, keep_vba=True)
    ws_atual = wb_atual["Geral"]
    
    for index, row in df_atual.iterrows():
        if pd.isna(row['AV']):
            id_atual = row['A']
            linha_base_dados = df_base_dados[df_base_dados['A'] == id_atual]

            if not linha_base_dados.empty:
                for i, col in enumerate(['E', 'F', 'G', 'H', 'I', 'J', 'K']):
                    ws_atual.cell(row=index+2, column=48+i).value = linha_base_dados.iloc[0][col]
    
    wb_atual.save(caminho_arquivo_atual)

preencher_celulas_vazias()
